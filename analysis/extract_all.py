# -*- coding: utf-8 -*-
"""
모든 거래내역 (IBK + 전자거래 + 카카오뱅크) 통합 파싱
"""
import os, sys, re, io
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
WORK = r"D:\Claw\workspace\analysis"

# === 1. IBK 거래내역조회 (HTML wrapped in .xls) ===
def parse_ibk_html(path):
    with open(path, "rb") as f:
        raw = f.read()
    text = raw.decode("utf-8", errors="replace")
    text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL|re.IGNORECASE)
    soup = BeautifulSoup(text, "lxml")
    meta = {}
    tables = soup.find_all("table")
    for tbl in tables:
        if "type1" in (tbl.get("class") or []):
            ths = [th.get_text(strip=True) for th in tbl.find_all("th")]
            tds = [td.get_text(strip=True) for td in tbl.find_all("td")]
            for h, d in zip(ths, tds):
                meta[h] = d
            break
    headers = None
    rows = []
    for tbl in tables:
        if "type2" in (tbl.get("class") or []):
            for tr in tbl.find_all("tr"):
                cells = [c.get_text(strip=True) for c in tr.find_all(["th","td"])]
                if not cells: continue
                if headers is None and any("거래일시" in v for v in cells):
                    headers = cells
                    continue
                if headers and len(cells) == len(headers):
                    rows.append(cells)
            break
    df = pd.DataFrame(rows, columns=headers) if headers else pd.DataFrame()
    return meta, df

# === 2. 카카오뱅크 .xlsx 파싱 ===
def parse_kakao_xlsx(path):
    print(f"\n--- KAKAO file: {os.path.basename(path)} ---")
    # Try multiple ways
    try:
        # Read raw with openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        print(f"  Sheet: {ws.title}, rows: {len(all_rows)}, cols: {ws.max_column}")
        # Print first 25 rows for inspection
        for i, r in enumerate(all_rows[:25]):
            print(f"  [{i}]", r)
        return all_rows
    except Exception as e:
        print("  openpyxl error:", e)
        return None

# === 3. 전자거래신청 상세내역조회 .xls 파싱 ===
def parse_ej_xls(path):
    print(f"\n--- EJ file: {os.path.basename(path)} ---")
    with open(path, "rb") as f:
        head = f.read(64)
    print("  Header hex:", head[:16].hex())
    print("  Header ascii:", head[:32].decode('ascii', errors='replace'))
    # Try as HTML
    with open(path, "rb") as f:
        raw = f.read()
    is_html = raw[:32].lower().startswith(b'<html') or b'<table' in raw[:1024].lower()
    if is_html:
        print("  Format: HTML")
        text = raw.decode("utf-8", errors="replace")
        # Print first 800 chars (after style)
        text_no_style = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL|re.IGNORECASE)
        soup = BeautifulSoup(text_no_style, "lxml")
        # All tables
        tables = soup.find_all("table")
        print(f"  Tables found: {len(tables)}")
        for ti, tbl in enumerate(tables):
            rows = tbl.find_all("tr")
            if not rows: continue
            print(f"  -- Table {ti} ({len(rows)} rows) class={tbl.get('class')} --")
            for ri, tr in enumerate(rows[:8]):
                cells = [c.get_text(strip=True) for c in tr.find_all(["th","td"])]
                print(f"    [{ri}]", cells)
        return tables
    return None

# === MAIN ===
print("="*100)
print("【모든 파일 구조 분석】")
print("="*100)

import glob
all_files = sorted(glob.glob(os.path.join(WORK, "*.xls")) + glob.glob(os.path.join(WORK, "*.xlsx")))
# 일관성 파일은 제외
all_files = [f for f in all_files if "일관성" not in os.path.basename(f)]

for f in all_files:
    name = os.path.basename(f)
    print("\n" + "="*100)
    print("FILE:", name)
    if name.startswith("농협"):
        parse_kakao_xlsx(f)
    elif name.startswith("과거거래"):
        parse_ej_xls(f)
    elif name.startswith("거래내역조회"):
        meta, df = parse_ibk_html(f)
        print("  META:", meta)
        print("  Rows:", len(df))
